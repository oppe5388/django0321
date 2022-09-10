from django.core.management.base import BaseCommand
from myinfo.models import *
import openpyxl

import time

class Command(BaseCommand):
    help = "CAエクセルインポート"

    def handle(self, *args, **options):

        # 時間計測開始
        time_sta = time.time()

        #ファイルは手動で配置
        wb = openpyxl.load_workbook('/code/CAs-2022-09-03.xlsx')
        ws = wb.worksheets[0]
        #シートの最終行を取得
        Sheet_Max_Row = ws.max_row

        add_cas = []  # bulk_create用　作成クエリの配列

        # 行をループ
        for j in range(2,Sheet_Max_Row + 1):#2行目から
            # #1件ずつのレコード作成
            # CAs.objects.create(
            #     id = int(ws.cell(row=j, column=1).value),#全データ削除してあればid指定してOK
            #     dealer = Dealers.objects.get(pk=int(ws.cell(row=j, column=2).value)),#ForeignKeyフィールド
            #     shop = ws.cell(row=j, column=3).value,
            #     shopcode = ws.cell(row=j, column=4).value,
            #     cacode = ws.cell(row=j, column=5).value,
            #     name = ws.cell(row=j, column=6).value,
            #     kana = ws.cell(row=j, column=7).value,
            # )
            # print(j)

            # bulk_createバージョン
            ca = CAs(
                id = int(ws.cell(row=j, column=1).value),#全データ削除してあればid指定してOK
                dealer = Dealers.objects.get(pk=int(ws.cell(row=j, column=2).value)),#ForeignKeyフィールド
                shop = ws.cell(row=j, column=3).value,
                shopcode = ws.cell(row=j, column=4).value,
                cacode = ws.cell(row=j, column=5).value,
                name = ws.cell(row=j, column=6).value,
                kana = ws.cell(row=j, column=7).value,
            )
            add_cas.append(ca)  # 作成クエリを配列に格納
            print(j)

        CAs.objects.bulk_create(add_cas)  # ここで追加のクエリ1つ発行

        # 時間計測終了
        time_end = time.time()
        # 経過時間（秒）
        tim = time_end- time_sta
        print(tim)

    # 引数を使う場合は必要
    def add_arguments(self, parser):
        parser.add_argument('--name', nargs='?', default='', type=str)